import openpyxl
import requests
import os
import datetime
import sys

# ================= 配置区 =================
# 默认寻找的文件名
TARGET_FILENAME = "专区信息汇总表.xlsx"
# API 路径
API_SUFFIX = "json/ZiZhanIndexCount.json"


# ==========================================

def log(msg):
    now = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] {msg}")


def process_logic():
    print("=" * 60)
    print("      BiddingCount 脚本执行模式 (直接运行版)")
    print("=" * 60)

    # 1. 自动定位文件路径 (脚本所在目录)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, TARGET_FILENAME)

    if not os.path.exists(file_path):
        log(f"错误: 在脚本目录下未找到 {TARGET_FILENAME}")
        return

    try:
        log(f"加载文件: {TARGET_FILENAME}")
        # 原地修改模式，不带 data_only 以保护原始公式
        wb = openpyxl.load_workbook(file_path)

        for sheet in wb.worksheets:
            if sheet.max_row < 2: continue
            log(f"正在扫描分表: {sheet.title}")

            # --- 寻列逻辑 ---
            addr_col, jy_col, kb_col = None, None, None
            last_filled_col = 0

            # 扫描表头
            for col in range(1, sheet.max_column + 1):
                cell_val = sheet.cell(row=1, column=col).value
                if cell_val:
                    last_filled_col = col
                    header_str = str(cell_val).replace(" ", "")
                    if "专区地址" in header_str: addr_col = col
                    if "近期交易" in header_str: jy_col = col
                    if "今日开标" in header_str: kb_col = col

            if not addr_col:
                log("跳过: 未发现‘专区地址’列")
                continue

            # --- 定位增量更新列 ---
            # 如果不存在结果列，则在最后一个有内容的列后面追加
            if not jy_col:
                jy_col = last_filled_col + 1
                kb_col = last_filled_col + 2
                sheet.cell(row=1, column=jy_col).value = "近期交易"
                sheet.cell(row=1, column=kb_col).value = "今日开标"
                log(f"新增列位置: {jy_col}, {kb_col}")
            else:
                if not kb_col: kb_col = jy_col + 1
                log(f"覆盖现有列位置: {jy_col}, {kb_col}")

            # --- 执行请求 ---
            success_count = 0
            for row in range(2, sheet.max_row + 1):
                raw_url = sheet.cell(row=row, column=addr_col).value
                if not raw_url or "http" not in str(raw_url):
                    continue

                # 自动处理空格和斜杠补全
                base_url = str(raw_url).strip()
                if not base_url.endswith('/'): base_url += '/'
                full_api_url = f"{base_url}{API_SUFFIX}"

                try:
                    resp = requests.get(full_api_url, timeout=5)
                    if resp.status_code == 200:
                        data = resp.json()
                        sheet.cell(row=row, column=jy_col).value = data.get("countjy", 0)
                        sheet.cell(row=row, column=kb_col).value = data.get("countkb", 0)
                        success_count += 1
                    else:
                        sheet.cell(row=row, column=jy_col).value = f"HTTP {resp.status_code}"
                except:
                    sheet.cell(row=row, column=jy_col).value = "请求超时"

            log(f"分表 [{sheet.title}] 完成，更新数: {success_count}")

        # 保存结果
        try:
            wb.save(file_path)
            log("SUCCESS: 数据已原地更新并保存。")
        except PermissionError:
            log("ERROR: 保存失败！请关闭 Excel 文件后再运行脚本。")

    except Exception as e:
        log(f"致命错误: {str(e)}")


if __name__ == "__main__":
    process_logic()
    print("=" * 60)
    # 如果是双击运行，保留窗口查看结果
    input("执行结束，按回车键退出...")