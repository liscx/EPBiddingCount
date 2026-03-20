import openpyxl
import requests
import os
import datetime


def process_excel(in_path, out_dir, log_func):
    try:
        log_func(f"正在读取: {os.path.basename(in_path)}")
        wb = openpyxl.load_workbook(in_path)
        api_path = "json/ZiZhanIndexCount.json"

        # --- 路径决策逻辑 ---
        in_dir = os.path.dirname(os.path.abspath(in_path))
        target_dir = os.path.abspath(out_dir)
        save_path = in_path if in_dir == target_dir else os.path.join(out_dir,
                                                                      f"回填结果_{datetime.datetime.now().strftime('%H%M%S')}.xlsx")

        log_func(f"保存模式：{'原地覆盖' if in_dir == target_dir else '另存为'}")

        for sheet in wb.worksheets:
            if sheet.max_row < 2: continue
            log_func(f"--- 扫描分表: {sheet.title} ---")

            # --- 1. 寻找现有列 ---
            addr_col, jy_col, kb_col = None, None, None
            last_filled_col = 0  # 记录有内容的最后一列

            # 遍历当前已有的所有列，寻找表头
            for col in range(1, sheet.max_column + 1):
                cell_val = sheet.cell(row=1, column=col).value
                if cell_val:
                    last_filled_col = col  # 只要有字，就更新最后有字列的位置
                    header_str = str(cell_val).replace(" ", "")
                    if "专区地址" in header_str: addr_col = col
                    if "近期交易" in header_str: jy_col = col
                    if "今日开标" in header_str: kb_col = col

            if not addr_col:
                log_func("跳过：未找到‘专区地址’列")
                continue

            # --- 2. 确定写入列的位置 ---
            # 如果没找到已有的“近期交易”列，就在“最后有字列”的后面开始写
            if not jy_col:
                jy_col = last_filled_col + 1
                kb_col = last_filled_col + 2
                sheet.cell(row=1, column=jy_col).value = "近期交易"
                sheet.cell(row=1, column=kb_col).value = "今日开标"
                log_func(f"在数据边缘新建列: 第 {jy_col} 和 {kb_col} 列")
            else:
                # 如果已经存在，kb_col 通常紧随其后或通过刚才的扫描已确定
                if not kb_col: kb_col = jy_col + 1  # 兜底逻辑
                log_func(f"匹配并覆盖现有列: 第 {jy_col} 和 {kb_col} 列")

            # --- 3. 执行 API 请求与数据回填 ---
            success_count = 0
            for row in range(2, sheet.max_row + 1):
                url_val = sheet.cell(row=row, column=addr_col).value
                if not url_val or "http" not in str(url_val): continue

                full_url = str(url_val).strip()
                if not full_url.endswith('/'): full_url += '/'

                try:
                    resp = requests.get(f"{full_url}{api_path}", timeout=5)
                    if resp.status_code == 200:
                        data = resp.json()
                        sheet.cell(row=row, column=jy_col).value = data.get("countjy", 0)
                        sheet.cell(row=row, column=kb_col).value = data.get("countkb", 0)
                        success_count += 1
                    else:
                        sheet.cell(row=row, column=jy_col).value = f"HTTP:{resp.status_code}"
                except:
                    sheet.cell(row=row, column=jy_col).value = "超时"

            log_func(f"分表 [{sheet.title}] 更新完成 ({success_count}条)")

        # --- 4. 保存 ---
        try:
            wb.save(save_path)
            log_func(f"\n[成功] 任务执行完毕！")
            return save_path
        except PermissionError:
            log_func("\n[错误] 保存失败！请确保 Excel 文件已关闭。")
            return None
    except Exception as e:
        log_func(f"运行出错: {str(e)}")
        return None